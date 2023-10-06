from tkinter import *
from openpyxl import *
from tkinter import ttk
from tkinter import messagebox


wb = load_workbook(r'C:\Users\Admin\Desktop\dangki.xlsx')
sheet = wb.active

def validate_student_id():
    numbers = numbers_field.get()
    if not numbers.isdigit() or len(numbers_field) != 7:
        messagebox.showerror("Lỗi", "Mã số sinh viên phải là 7 chữ số.")
    else:
        return True

def excel():
	

	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50

 

	sheet.cell(row=1, column=1).value = "Mssv"
	sheet.cell(row=1, column=2).value = "Họ tên"
	sheet.cell(row=1, column=3).value = "Ngày sinh"
	sheet.cell(row=1, column=4).value = "Email"
	sheet.cell(row=1, column=5).value = "Số điện thoại"
	sheet.cell(row=1, column=6).value = "Học kỳ "
	sheet.cell(row=1, column=7).value = "Năm học"
  

def focus1(event):
	numbers_field.focus_set()

def focus2(event):
	name_field.focus_set()

def focus3(event):
	daytime_field.focus_set()

def focus4(event):
	email_field.focus_set()
 
def focus5(event):
	phone_field.focus_set()


def focus6(event):
	semester_field.focus_set()
 
def focus7(event):
	year_field.focus_set()
 


def clear():
	
	numbers_field.delete(0, END)
	name_field.delete(0, END)
	daytime_field.delete(0, END)
	email_field.delete(0, END)
	phone_field.delete(0, END)
	semester_field.delete(0, END)
	year_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
	
	# if user not fill any entry
	# then print "empty input"
	if (numbers_field.get() == "" and
		name_field.get() == "" and
		daytime_field.get() == "" and
		email_field.get() == "" and
		phone_field.get() == "" and
		semester_field.get() == "" and
		year_field.get() == ""):

        
			
		print("empty input")

	else:

		current_row = sheet.max_row
		current_column = sheet.max_column

		sheet.cell(row=current_row + 1, column=1).value = numbers_field.get()
		sheet.cell(row=current_row + 1, column=2).value = name_field.get()
		sheet.cell(row=current_row + 1, column=3).value = daytime_field.get()
		sheet.cell(row=current_row + 1, column=4).value = email_field.get()
		sheet.cell(row=current_row + 1, column=5).value = phone_field.get()
		sheet.cell(row=current_row + 1, column=6).value = semester_field.get()
		sheet.cell(row=current_row + 1, column=7).value = year_field.get()

		# save the file
		wb.save(r'C:\Users\Admin\Desktop\dangki.xlsx')

		# set focus on the name_field box
		name_field.focus_set()

		# call the clear() function
		clear()

  
if __name__ == "__main__":
    root = Tk()
    root.configure(background='light blue')
    root.title("Đăng kí học phần")
    root.geometry("550x300")
    #tao Form label
    heading = Label(root, text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN", font=('Serif 14'), fg="Red" , bg="light blue")
    numbers = Label(root, text="Mã số sinh viên", bg="light blue")
    fullname = Label(root, text="Họ tên", bg="light blue")
    daytime = Label(root, text="Ngày sinh", bg="light blue")
    email = Label(root, text="Email", bg="light blue")
    phone = Label(root, text="Số điện thoại", bg="light blue")
    semester = Label(root, text="Học kỳ", bg="light blue")
    year = Label(root, text="Năm học", bg="light blue")
    choose_subjects = Label(root, text="Chọn môn học", bg="light blue")
    cb_1 = Checkbutton(root, text="Lập trình Python", bg="light blue")
    cb_2 = Checkbutton(root, text="Lập trình Công nghệ phần mềm", bg="light blue")
    cb_3 = Checkbutton(root, text="Lập trình Java", bg="light blue")
    cb_4 = Checkbutton(root, text="Lập trình Phát triển ứng dụng web", bg="light blue")
    btn_singup =Button(root, text="Đăng ký", bg="green1", command=insert)
    btn_exit = Button(root, text="Thoát", bg="green1", command=root.quit)
   
    #hien box
    heading.grid(row=0, column=1)
    numbers.grid(row=1, column=0)
    fullname.grid(row=2, column=0)
    daytime.grid(row=3, column=0)
    email.grid(row=4, column=0)
    phone.grid(row=5, column=0)
    semester.grid(row=6, column=0)
    year.grid(row=7, column=0)
    btn_singup.grid(row=11, column=0)
    btn_exit.grid(row=11, column=1)
  
    choose_subjects.grid(row=8, column=0)
    cb_1.grid(row=9, column=0)
    cb_2.grid(row=10, column=1)
    cb_3.grid(row=9, column=0)
    cb_4.grid(row=10, column=1)
   
    btn_exit.grid(row=11, column=1)
    
    
   
    # tao box
    numbers_field = Entry(root)
    name_field = Entry(root)
    daytime_field = Entry(root)
    email_field = Entry(root)
    phone_field = Entry(root)
    semester_field = Entry(root)
    year_field = ttk.Combobox(root)
    
    #
    numbers_field.bind("<Return>", focus1)
    name_field.bind("<Return>", focus2)
    daytime_field.bind("<Return>", focus3)
    email_field.bind("<Return>", focus4)
    phone_field.bind("<Return>", focus5)
    semester_field.bind("<Return>", focus6)
    year_field.bind("<Return>", focus7)
    
    #ds nam hoc trong combobox
    year_field['values'] =('2022-2023', '2023-2024','2024-2025')
    year_field.current()
    #chieu dai hop
    numbers_field.grid(row=1, column=1, ipadx="100")
    name_field.grid(row=2, column=1, ipadx="100")
    daytime_field.grid(row=3, column=1, ipadx="100")
    email_field.grid(row=4, column=1, ipadx="100")
    phone_field.grid(row=5, column=1, ipadx="100")
    semester_field.grid(row=6, column=1, ipadx="100")
    year_field.grid(row=7, column=1, ipadx="92")
    cb_1.grid(row=9, column=1, ipadx='100')
    #goi ham 
    excel()
    #
    
    
    
    def onExit(self):
        self.quit()
   #
    root.mainloop()